"""Importación de clientes desde Excel (.xlsx) o CSV."""

import csv
import io
from typing import Any

from django.db import transaction

from cotizaciones.models import Cliente

# Encabezados aceptados (minúsculas, sin acentos opcionales)
COLUMN_ALIASES = {
    "nombre": ("nombre", "cliente", "razon social", "razón social", "name"),
    "email": ("email", "correo", "mail", "e-mail"),
    "telefono": ("telefono", "teléfono", "tel", "telefono", "phone"),
    "direccion": ("direccion", "dirección", "domicilio", "address"),
    "localidad": ("localidad", "ciudad", "city"),
    "activo": ("activo", "estado", "active", "habilitado"),
}


def _normalize_header(value: str) -> str:
    return (value or "").strip().lower()


def _map_headers(row: list[str]) -> dict[str, int]:
    """Devuelve {campo_modelo: índice_columna}."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(row):
        h = _normalize_header(str(raw) if raw is not None else "")
        if not h:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if h in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _parse_activo(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return True
    s = str(value).strip().lower()
    if s in ("1", "true", "si", "sí", "yes", "activo", "s", "y"):
        return True
    if s in ("0", "false", "no", "inactivo", "n"):
        return False
    return True


def _cell(row: list, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _upsert_cliente(data: dict[str, Any]) -> tuple[str, Cliente]:
    nombre = (data.get("nombre") or "").strip()
    if len(nombre) < 2:
        raise ValueError("Nombre inválido o vacío")

    email = (data.get("email") or "").strip() or None
    defaults = {
        "nombre": nombre,
        "telefono": (data.get("telefono") or "").strip() or None,
        "direccion": (data.get("direccion") or "").strip() or None,
        "localidad": (data.get("localidad") or "").strip() or None,
        "activo": data.get("activo", True),
    }

    if email:
        obj, created = Cliente.objects.update_or_create(
            email=email,
            defaults=defaults,
        )
        return ("creado" if created else "actualizado", obj)

    existente = Cliente.objects.filter(nombre__iexact=nombre).first()
    if existente:
        for key, val in defaults.items():
            setattr(existente, key, val)
        existente.save()
        return ("actualizado", existente)

    obj = Cliente.objects.create(**defaults)
    return ("creado", obj)


def importar_clientes_desde_archivo(archivo) -> dict:
    """
    Importa filas desde .xlsx o .csv.
    Retorna: {creados, actualizados, errores: [(fila, mensaje), ...]}
    """
    nombre = (archivo.name or "").lower()
    if nombre.endswith(".csv"):
        rows = _read_csv(archivo)
    elif nombre.endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx(archivo)
    else:
        raise ValueError("Formato no soportado. Usá archivo .xlsx o .csv")

    if not rows:
        raise ValueError("El archivo está vacío.")

    header = rows[0]
    col_map = _map_headers(header)
    if "nombre" not in col_map:
        raise ValueError(
            'No se encontró la columna "nombre". '
            "La primera fila debe incluir encabezados como: nombre, email, teléfono, localidad."
        )

    resultado = {"creados": 0, "actualizados": 0, "errores": []}

    with transaction.atomic():
        for num, row in enumerate(rows[1:], start=2):
            try:
                data = {
                    "nombre": _cell(row, col_map.get("nombre")),
                    "email": _cell(row, col_map.get("email")),
                    "telefono": _cell(row, col_map.get("telefono")),
                    "direccion": _cell(row, col_map.get("direccion")),
                    "localidad": _cell(row, col_map.get("localidad")),
                    "activo": _parse_activo(
                        row[col_map["activo"]] if "activo" in col_map and col_map["activo"] < len(row) else None
                    ),
                }
                if not data["nombre"]:
                    continue
                accion, _ = _upsert_cliente(data)
                if accion == "creado":
                    resultado["creados"] += 1
                else:
                    resultado["actualizados"] += 1
            except Exception as exc:
                resultado["errores"].append((num, str(exc)))

    return resultado


def _read_csv(archivo) -> list[list[str]]:
    raw = archivo.read()
    if isinstance(raw, bytes):
        for enc in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")
    else:
        text = raw
    reader = csv.reader(io.StringIO(text))
    return [list(r) for r in reader]


def _read_xlsx(archivo) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Falta la librería openpyxl para leer Excel. "
            "Contactá al administrador del sistema."
        ) from exc

    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([c if c is not None else "" for c in row])
    wb.close()
    return rows
