"""Exportación de clientes a Excel y PDF."""

from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from django.db.models import Q

from cotizaciones.models import Cliente


def clientes_queryset_filtrado(request):
    """Mismos filtros que la lista de clientes."""
    qs = Cliente.objects.all().order_by("nombre")
    search = request.GET.get("search", "").strip()
    if search:
        qs = qs.filter(
            Q(nombre__icontains=search)
            | Q(telefono__icontains=search)
            | Q(localidad__icontains=search)
            | Q(email__icontains=search)
        )
    estado = request.GET.get("estado", "activos")
    if estado == "activos":
        qs = qs.filter(activo=True)
    elif estado == "inactivos":
        qs = qs.filter(activo=False)
    return qs


def exportar_clientes_excel_response(clientes) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["Nombre", "Email", "Teléfono", "Dirección", "Localidad", "Estado"]
    header_fill = PatternFill("solid", fgColor="1C3A5E")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, c in enumerate(clientes, start=2):
        ws.cell(row=row_idx, column=1, value=c.nombre)
        ws.cell(row=row_idx, column=2, value=c.email or "")
        ws.cell(row=row_idx, column=3, value=c.telefono or "")
        ws.cell(row=row_idx, column=4, value=c.direccion or "")
        ws.cell(row=row_idx, column=5, value=c.localidad or "")
        ws.cell(row=row_idx, column=6, value="Activo" if c.activo else "Inactivo")

    widths = [32, 28, 16, 36, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = datetime.now().strftime("%Y%m%d")
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="clientes_{fecha}.xlsx"'
    return response


def exportar_clientes_pdf_response(clientes) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=28,
        leftMargin=28,
        topMargin=36,
        bottomMargin=28,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TituloClientes",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1C3A5E"),
        spaceAfter=12,
    )
    fecha_txt = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements = [
        Paragraph("Listado de clientes — GCinsumos", title_style),
        Paragraph(f"<font size=9 color='#64748b'>Generado: {fecha_txt} · Total: {len(clientes)}</font>", styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [["Nombre", "Email", "Teléfono", "Localidad", "Estado"]]
    for c in clientes:
        data.append([
            c.nombre[:40],
            (c.email or "—")[:35],
            c.telefono or "—",
            (c.localidad or "—")[:25],
            "Activo" if c.activo else "Inactivo",
        ])

    if len(data) == 1:
        data.append(["Sin clientes con los filtros aplicados", "", "", "", ""])

    table = Table(
        data,
        colWidths=[2.8 * inch, 2.4 * inch, 1.2 * inch, 1.4 * inch, 0.9 * inch],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C3A5E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF3F9")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#AACAE6")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AACAE6")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    fecha = datetime.now().strftime("%Y%m%d")
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="clientes_{fecha}.pdf"'
    return response
