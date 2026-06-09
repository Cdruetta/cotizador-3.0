import csv
import io
from typing import Any

from django.db import transaction
from decimal import Decimal, InvalidOperation

from cotizaciones.models import Producto, Proveedor


COLUMN_ALIASES = {
    "nombre": ("nombre", "producto", "name", "product"),
    "tipo": ("tipo", "type"),
    "proveedor": ("proveedor", "provider", "supplier"),
    "stock": ("stock", "cantidad", "quantity"),
    "precio_unitario": ("precio", "precio unitario", "precio_unitario", "price", "cost"),
    "activo": ("activo", "estado", "active", "habilitado"),
}

TIPO_MAP = {
    "producto": "producto",
    "servicio_soft": "servicio_soft",
    "servicio_hard": "servicio_hard",
    "servicio (software)": "servicio_soft",
    "servicio (hardware)": "servicio_hard",
    "software": "servicio_soft",
    "hardware": "servicio_hard",
    "servicio": "servicio_soft",
}


def _normalize_header(value: str) -> str:
    return (value or "").strip().lower()


def _map_headers(row: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(row):
        h = _normalize_header(str(raw) if raw is not None else "")
        if not h:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if h in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _parse_activo(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return True
    s = str(value).strip().lower()
    return s not in ("0", "false", "no", "inactivo", "n")


def _parse_tipo(value: Any) -> str:
    if value is None:
        return "producto"
    s = str(value).strip().lower()
    return TIPO_MAP.get(s, "producto")


def _parse_precio(value: Any) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0.00")
    s = str(value).strip().replace("$", "").replace(",", ".").strip()
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0.00")


def _parse_stock(value: Any) -> int:
    if value is None:
        return 0
    s = str(value).strip()
    try:
        return max(0, int(float(s)))
    except (ValueError, InvalidOperation):
        return 0


def _cell(row: list, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _upsert_producto(data: dict[str, Any]) -> tuple[str, Producto]:
    nombre = (data.get("nombre") or "").strip()
    if len(nombre) < 2:
        raise ValueError("Nombre inválido o vacío")

    proveedor_nombre = (data.get("proveedor") or "").strip()
    proveedor = None
    if proveedor_nombre:
        proveedor = Proveedor.objects.filter(nombre__iexact=proveedor_nombre).first()
        if not proveedor:
            proveedor = Proveedor.objects.create(nombre=proveedor_nombre)

    defaults = {
        "nombre": nombre,
        "tipo": data.get("tipo", "producto"),
        "stock": data.get("stock", 0),
        "precio_unitario": data.get("precio_unitario", Decimal("0.00")),
        "activo": data.get("activo", True),
    }
    if proveedor:
        defaults["proveedor"] = proveedor

    existente = Producto.objects.filter(nombre__iexact=nombre).first()
    if existente:
        for key, val in defaults.items():
            setattr(existente, key, val)
        if proveedor:
            existente.proveedor = proveedor
        existente.save()
        return ("actualizados", existente)

    if not proveedor:
        raise ValueError("Se requiere proveedor para crear un producto nuevo")
    obj = Producto.objects.create(**defaults)
    return ("creados", obj)


def importar_productos_desde_archivo(archivo) -> dict:
    nombre = (archivo.name or "").lower()
    if nombre.endswith(".csv"):
        rows = _read_csv(archivo)
    elif nombre.endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx(archivo)
    else:
        raise ValueError("Formato no soportado. Usá archivo .xlsx o .csv")

    if not rows:
        raise ValueError("El archivo está vacío.")

    header = rows[0]
    col_map = _map_headers(header)
    if "nombre" not in col_map:
        raise ValueError(
            'No se encontró la columna "nombre". '
            "La primera fila debe incluir encabezados como: nombre, tipo, proveedor, stock, precio."
        )

    resultado = {"creados": 0, "actualizados": 0, "errores": []}

    with transaction.atomic():
        for num, row in enumerate(rows[1:], start=2):
            try:
                data = {
                    "nombre": _cell(row, col_map.get("nombre")),
                    "tipo": _parse_tipo(_cell(row, col_map.get("tipo"))),
                    "proveedor": _cell(row, col_map.get("proveedor")),
                    "stock": _parse_stock(_cell(row, col_map.get("stock"))),
                    "precio_unitario": _parse_precio(_cell(row, col_map.get("precio_unitario"))),
                    "activo": _parse_activo(
                        row[col_map["activo"]] if "activo" in col_map and col_map["activo"] < len(row) else None
                    ),
                }
                if not data["nombre"]:
                    continue
                accion, _ = _upsert_producto(data)
                resultado[accion] += 1
            except Exception as exc:
                resultado["errores"].append((num, str(exc)))

    return resultado


def _read_csv(archivo) -> list[list[str]]:
    raw = archivo.read()
    if isinstance(raw, bytes):
        for enc in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")
    else:
        text = raw
    reader = csv.reader(io.StringIO(text))
    return [list(r) for r in reader]


def _read_xlsx(archivo) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Falta la librería openpyxl para leer Excel. "
            "Contactá al administrador del sistema."
        ) from exc

    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([c if c is not None else "" for c in row])
    wb.close()
    return rows
