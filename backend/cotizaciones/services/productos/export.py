from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from django.db.models import Q

from cotizaciones.models import Producto


def productos_queryset_filtrado(request):
    qs = Producto.objects.select_related("proveedor").all().order_by("nombre")
    q = request.GET.get("q", "").strip()
    proveedor_id = request.GET.get("proveedor")
    activo = request.GET.get("activo")
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    if proveedor_id:
        qs = qs.filter(proveedor_id=proveedor_id)
    if activo == "1":
        qs = qs.filter(activo=True)
    elif activo == "0":
        qs = qs.filter(activo=False)
    return qs


def exportar_productos_excel_response(productos) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = ["Producto", "Tipo", "Proveedor", "Stock", "Precio", "Estado"]
    header_fill = PatternFill("solid", fgColor="1C3A5E")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, p in enumerate(productos, start=2):
        ws.cell(row=row_idx, column=1, value=p.nombre)
        ws.cell(row=row_idx, column=2, value=p.get_tipo_display())
        ws.cell(row=row_idx, column=3, value=p.proveedor.nombre if p.proveedor else "")
        ws.cell(row=row_idx, column=4, value=p.stock)
        ws.cell(row=row_idx, column=5, value=float(p.precio_unitario))
        ws.cell(row=row_idx, column=6, value="Activo" if p.activo else "Inactivo")

    widths = [36, 18, 22, 10, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = datetime.now().strftime("%Y%m%d")
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="stock_{fecha}.xlsx"'
    return response


def exportar_productos_pdf_response(productos) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=28, leftMargin=28, topMargin=36, bottomMargin=28,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TituloStock",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1C3A5E"),
        spaceAfter=12,
    )
    fecha_txt = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements = [
        Paragraph("Listado de stock — GCinsumos", title_style),
        Paragraph(f"<font size=9 color='#64748b'>Generado: {fecha_txt} · Total: {len(productos)}</font>", styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [["Producto", "Tipo", "Proveedor", "Stock", "Precio", "Estado"]]
    for p in productos:
        data.append([
            p.nombre[:45],
            p.get_tipo_display(),
            p.proveedor.nombre[:25] if p.proveedor else "—",
            str(p.stock),
            f"${float(p.precio_unitario):.2f}",
            "Activo" if p.activo else "Inactivo",
        ])

    if len(data) == 1:
        data.append(["Sin productos con los filtros aplicados", "", "", "", "", ""])

    table = Table(
        data,
        colWidths=[2.8 * inch, 1.4 * inch, 1.8 * inch, 0.7 * inch, 0.9 * inch, 0.8 * inch],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C3A5E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF3F9")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#AACAE6")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AACAE6")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    fecha = datetime.now().strftime("%Y%m%d")
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="stock_{fecha}.pdf"'
    return response
